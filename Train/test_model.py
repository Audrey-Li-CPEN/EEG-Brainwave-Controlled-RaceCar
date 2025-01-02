import numpy as np
import pandas as pd
import torch
from torch.utils.data import Dataset, DataLoader
from scipy.signal import butter, filtfilt
from sklearn.preprocessing import StandardScaler
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# DeepCNN Model
# DeepCNN Model
class DeepCNN1D(torch.nn.Module):
    def __init__(self, num_channels=14, num_classes=4):
        super(DeepCNN1D, self).__init__()
        self.conv1 = torch.nn.Conv1d(num_channels, 64, kernel_size=31, padding=15)
        self.bn1 = torch.nn.BatchNorm1d(64)
        self.relu = torch.nn.ReLU()
        self.resblock1 = self._residual_block(64)
        self.resblock2 = self._residual_block(64)
        self.global_pool = torch.nn.AdaptiveAvgPool1d(1)
        self.fc = torch.nn.Sequential(
            torch.nn.Linear(64, 32),
            torch.nn.ReLU(),
            torch.nn.Dropout(0.5),
            torch.nn.Linear(32, num_classes)
        )

    def _residual_block(self, channels):
        return torch.nn.Sequential(
            torch.nn.Conv1d(channels, channels, kernel_size=15, padding=7),
            torch.nn.BatchNorm1d(channels),
            torch.nn.ReLU(),
            torch.nn.Conv1d(channels, channels, kernel_size=15, padding=7),
            torch.nn.BatchNorm1d(channels),
            torch.nn.ReLU()
        )

    def forward(self, x):
        x = self.conv1(x)
        x = self.bn1(x)
        x = self.relu(x)
        x = self.resblock1(x)
        x = self.resblock2(x)
        x = self.global_pool(x).squeeze(-1)
        x = self.fc(x)
        return x

# Dataset class for testing
class EEGDataset(Dataset):
    def __init__(self, features):
        self.features = torch.tensor(features, dtype=torch.float32)

    def __len__(self):
        return len(self.features)

    def __getitem__(self, idx):
        return self.features[idx]

# Preprocess the test data
def preprocess_test_data(file_path, label_mapping, fs=128, lowcut=8.0, highcut=50.0, window_size=50, overlap=45):
    data = pd.read_csv(file_path)
    print("Columns in CSV:", data.columns)

    # Identify EEG channels
    eeg_channels = [col for col in data.columns if col.startswith('Channel')]

    # Apply bandpass filter
    def butter_bandpass(lowcut, highcut, fs, order=4):
        nyquist = 0.5 * fs
        low = lowcut / nyquist
        high = highcut / nyquist
        return butter(order, [low, high], btype='band')

    def butter_bandpass_filter(data, lowcut, highcut, fs, order=4):
        b, a = butter_bandpass(lowcut, highcut, fs, order=order)
        return filtfilt(b, a, data)

    for col in eeg_channels:
        data[col] = butter_bandpass_filter(data[col].values, lowcut, highcut, fs)

    # Sliding window
    features = []
    indices = []  # To track which rows are covered by each window
    step_size = window_size - overlap
    for start in range(0, len(data) - window_size + 1, step_size):
        end = start + window_size
        window = data.iloc[start:end][eeg_channels].values.T
        features.append(window)
        indices.append((start, end))

    features = np.array(features)

    # Normalize features
    scaler = StandardScaler()
    for i in range(features.shape[1]):
        features[:, i, :] = scaler.fit_transform(features[:, i, :])

    return features, indices, data

# Prediction function
def predict(file_path, model_path, label_mapping, output_path, window_size=50, overlap=45):
    features, indices, original_data = preprocess_test_data(file_path, label_mapping)

    # Load the model
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    model = DeepCNN1D(num_channels=14, num_classes=len(label_mapping)).to(device)
    model.load_state_dict(torch.load(model_path))
    model.eval()

    # Create dataset and dataloader
    test_dataset = EEGDataset(features)
    test_loader = DataLoader(test_dataset, batch_size=32, shuffle=False)

    # Make predictions
    predicted_tags = []
    with torch.no_grad():
        for inputs in tqdm(test_loader, desc="Predicting"):
            inputs = inputs.to(device)
            outputs = model(inputs)
            _, predicted = torch.max(outputs, 1)
            predicted_tags.extend(predicted.cpu().numpy())

    # Convert predicted labels back to tag names
    inverse_label_mapping = {v: k for k, v in label_mapping.items()}
    predicted_tags = [inverse_label_mapping[tag] for tag in predicted_tags]

    # Assign predictions to the original data
    tag_predict = [""] * len(original_data)  # Initialize with empty strings
    for (start, end), tag in zip(indices, predicted_tags):
        for i in range(start, end):
            if not tag_predict[i]:  # If the row is not yet assigned
                tag_predict[i] = tag

    original_data['Tag_predict'] = tag_predict

    # Create Excel file with colors
    wb = Workbook()
    ws = wb.active
    ws.title = "Predictions"

    # Define styles
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Write headers
    for col_num, column_title in enumerate(original_data.columns, start=1):
        ws.cell(row=1, column=col_num, value=column_title)

    # Write data with colors
    for row_num, row in original_data.iterrows():
        for col_num, value in enumerate(row, start=1):
            cell = ws.cell(row=row_num + 2, column=col_num, value=value)
            if col_num == len(original_data.columns):  # Tag_predict column
                if row['Tag'] == row['Tag_predict']:
                    cell.fill = green_fill
                else:
                    cell.fill = red_fill

    wb.save(output_path)
    print(f"Predictions saved to {output_path}")


if __name__ == "__main__":
    # Paths
    test_file_path = r"D:\Train\processed_test_with_tags4.csv"
    model_path = r"D:\Train\trained_model.pth"
    output_path = r"D:\Train\test_model_with_predictions.xlsx"

    # Define consistent label mapping
    label_mapping = {'left': 0, 'rest': 1, 'right': 2, 'stop': 3}

    # Predict and save results
    predict(test_file_path, model_path, label_mapping, output_path)
